"""
services.ingestion
====================
Idempotent document ingestion pipeline with MongoDB.

Pipeline steps
--------------
1. **Extract** raw text from PDF / DOCX / XLSX / TXT.
2. **Split** into overlapping chunks (1 000 chars, 200 overlap).
3. **Deduplicate** chunks by SHA-256 hash within the document.
4. **Embed** each chunk via ``services.embedding``.
5. **Store** chunks + embeddings in ``document_chunks`` collection.
6. **Record** the ingestion run in ``ingestion_manifests`` collection.

Idempotency is achieved by hashing the source file: if a document with
the same ``file_hash`` already exists and is ``completed``, re-ingestion
is skipped.
"""

import hashlib
import time
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from motor.motor_asyncio import AsyncIOMotorDatabase

from app.config import settings
from app.services.embedding import embed_texts


# ────────────────────────────────────────────────────────────────
# Text extraction helpers
# ────────────────────────────────────────────────────────────────

def _extract_pdf(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Extract text from a PDF file, one entry per page.

    Returns:
        List of dicts with keys ``text`` and ``source_page``.
    """
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(file_bytes))
    pages = []
    for i, page in enumerate(reader.pages):
        page_text = page.extract_text() or ""
        if page_text.strip():
            pages.append({"text": page_text, "source_page": str(i + 1)})
    return pages


def _extract_docx(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Extract text from a DOCX file.

    All paragraphs are concatenated into a single entry because DOCX
    doesn't have a natural page concept.
    """
    from docx import Document as DocxDocument

    doc = DocxDocument(BytesIO(file_bytes))
    full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if not full_text.strip():
        return []
    return [{"text": full_text, "source_page": None}]


def _extract_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Extract text from an Excel workbook, one entry per sheet.

    Each cell value is cast to a string and rows are joined by tabs;
    rows are joined by newlines.
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                rows.append(row_text)
        if rows:
            sheets.append({"text": "\n".join(rows), "source_page": sheet_name})
    return sheets


def _extract_txt(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Extract text from a plain-text file.

    Uses ``chardet`` to detect encoding if UTF-8 decoding fails.
    """
    try:
        text_content = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        import chardet
        detected = chardet.detect(file_bytes)
        text_content = file_bytes.decode(detected.get("encoding", "utf-8"), errors="replace")

    if not text_content.strip():
        return []
    return [{"text": text_content, "source_page": None}]


# Map of supported file types to their extraction functions.
EXTRACTORS = {
    "pdf": _extract_pdf,
    "docx": _extract_docx,
    "xlsx": _extract_xlsx,
    "txt": _extract_txt,
}


# ────────────────────────────────────────────────────────────────
# Chunking
# ────────────────────────────────────────────────────────────────

def _split_text(
    text_content: str,
    chunk_size: int = 1000,
    chunk_overlap: int = 200,
) -> list[str]:
    """
    Split text into overlapping chunks of roughly ``chunk_size`` characters.

    Uses a simple character-level sliding window.

    Args:
        text_content  : The full text to split.
        chunk_size    : Target size of each chunk in characters.
        chunk_overlap : Number of overlapping characters between consecutive chunks.

    Returns:
        List of chunk strings.
    """
    if len(text_content) <= chunk_size:
        return [text_content]

    chunks = []
    start = 0
    while start < len(text_content):
        end = start + chunk_size
        chunk = text_content[start:end]
        if chunk.strip():
            chunks.append(chunk.strip())
        start += chunk_size - chunk_overlap
    return chunks


# ────────────────────────────────────────────────────────────────
# Main ingestion function
# ────────────────────────────────────────────────────────────────

async def ingest_document(
    db: AsyncIOMotorDatabase,
    file_bytes: bytes,
    filename: str,
    user_id: uuid.UUID | None = None,
) -> dict:
    """
    Full ingestion pipeline: extract → chunk → embed → store.

    Args:
        db         : Async MongoDB database.
        file_bytes : Raw bytes of the uploaded file.
        filename   : Original filename (used to detect type).
        user_id    : Optional uploader's user ID.

    Returns:
        The Document dict (status will be ``completed`` or ``failed``).

    Raises:
        ValueError : If the file type is not supported.
    """
    start_time = time.time()

    # ── 1. Determine file type ──────────────────────────────────
    extension = filename.rsplit(".", maxsplit=1)[-1].lower() if "." in filename else ""
    if extension not in EXTRACTORS:
        raise ValueError(
            f"Unsupported file type: '.{extension}'. "
            f"Supported: {', '.join(EXTRACTORS.keys())}"
        )

    # ── 2. Compute file hash for deduplication ──────────────────
    file_hash = hashlib.sha256(file_bytes).hexdigest()

    # Check if this exact file has already been ingested.
    existing = await db.documents.find_one(
        {"file_hash": file_hash, "status": "completed"}
    )
    if existing:
        existing["id"] = existing["_id"]
        return existing  # idempotent — skip re-ingestion

    # ── 3. Create Document record ───────────────────────────────
    doc_id = uuid.uuid4()
    doc = {
        "_id": doc_id,
        "filename": filename,
        "file_type": extension,
        "file_hash": file_hash,
        "status": "processing",
        "uploaded_by": user_id,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
    }
    await db.documents.insert_one(doc)
    doc["id"] = doc_id

    try:
        # ── 4. Extract text ─────────────────────────────────────
        extractor = EXTRACTORS[extension]
        pages = extractor(file_bytes)
        if not pages:
            raise ValueError("No text could be extracted from the file.")

        # ── 5. Split into chunks ────────────────────────────────
        all_chunks: list[dict[str, Any]] = []
        for page_info in pages:
            raw_chunks = _split_text(page_info["text"])
            for chunk_text in raw_chunks:
                all_chunks.append({
                    "text": chunk_text,
                    "source_page": page_info["source_page"],
                })

        # ── 6. Deduplicate by content hash ──────────────────────
        seen_hashes: set[str] = set()
        unique_chunks: list[dict[str, Any]] = []
        for chunk in all_chunks:
            chunk_hash = hashlib.sha256(chunk["text"].encode()).hexdigest()
            if chunk_hash not in seen_hashes:
                seen_hashes.add(chunk_hash)
                chunk["hash"] = chunk_hash
                unique_chunks.append(chunk)

        # ── 7. Generate embeddings ──────────────────────────────
        chunk_texts = [c["text"] for c in unique_chunks]
        embeddings = await embed_texts(chunk_texts)

        # ── 8. Build MongoDB documents and store ────────────────
        chunk_objects = []
        for idx, (chunk_info, embedding) in enumerate(zip(unique_chunks, embeddings)):
            chunk_obj = {
                "_id": uuid.uuid4(),
                "document_id": doc_id,
                "chunk_index": idx,
                "content": chunk_info["text"],
                "embedding": embedding,
                "embedding_model": settings.EMBEDDING_MODEL,
                "source_page": chunk_info["source_page"],
                "chunk_hash": chunk_info["hash"],
                "created_at": datetime.now(timezone.utc),
            }
            chunk_objects.append(chunk_obj)

        if chunk_objects:
            await db.document_chunks.insert_many(chunk_objects)

        # ── 9. Mark document as completed ──────────────────────
        await db.documents.update_one(
            {"_id": doc_id},
            {
                "$set": {
                    "status": "completed",
                    "total_chunks": len(chunk_objects),
                    "updated_at": datetime.now(timezone.utc),
                }
            }
        )
        doc["status"] = "completed"
        doc["total_chunks"] = len(chunk_objects)

        # ── 10. Record ingestion manifest ───────────────────────
        elapsed = time.time() - start_time
        manifest = {
            "_id": uuid.uuid4(),
            "document_id": doc_id,
            "source_hash": file_hash,
            "chunk_count": len(chunk_objects),
            "embedding_model": settings.EMBEDDING_MODEL,
            "success": True,
            "processing_time_sec": round(elapsed, 2),
            "created_at": datetime.now(timezone.utc),
        }
        await db.ingestion_manifests.insert_one(manifest)

    except Exception as exc:
        # Mark the document as failed and record the error.
        await db.documents.update_one(
            {"_id": doc_id},
            {
                "$set": {
                    "status": "failed",
                    "error_message": str(exc),
                    "updated_at": datetime.now(timezone.utc),
                }
            }
        )
        doc["status"] = "failed"
        doc["error_message"] = str(exc)

        elapsed = time.time() - start_time
        manifest = {
            "_id": uuid.uuid4(),
            "document_id": doc_id,
            "source_hash": file_hash,
            "chunk_count": 0,
            "embedding_model": settings.EMBEDDING_MODEL,
            "success": False,
            "error_message": str(exc),
            "processing_time_sec": round(elapsed, 2),
            "created_at": datetime.now(timezone.utc),
        }
        await db.ingestion_manifests.insert_one(manifest)
        raise

    return doc
